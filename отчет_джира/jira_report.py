"""
Сбор баглистов (тип задачи «Баг») за выбранный месяц из Jira
и всех их сабтасков с выгрузкой в Excel.

Конфигурация: config.json (jira_base_url, jira_email, month, project).
Токен Jira: в .env (JIRA_API_TOKEN). См. .env.example.
Месяц в формате YYYY-MM (например 2025-01).
"""

import json
import os

from dotenv import load_dotenv

# Токен берётся из .env (не коммитится)
load_dotenv(os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env"))
import calendar
import re
from datetime import datetime, timedelta, time
from concurrent.futures import ThreadPoolExecutor, as_completed
from urllib.parse import urljoin

import requests
from requests.auth import HTTPBasicAuth
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# Тип задачи баглиста в Jira (русское название)
ISSUE_TYPE_BUGLIST = "Баг"
STATUS_TASK = "TASK"
STATUS_DEV_TEST = "DEV TEST"

WORKDAY_START = time(9, 0)
WORKDAY_LUNCH_START = time(12, 0)
WORKDAY_LUNCH_END = time(13, 0)
WORKDAY_END = time(18, 0)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_PATH = os.path.join(SCRIPT_DIR, "config.json")


def load_config():
    path = CONFIG_PATH
    if not os.path.exists(path):
        raise FileNotFoundError(
            f"Файл конфигурации не найден: {path}\n"
            "Создайте config.json с полями: jira_base_url, jira_email, month, project. "
            "Токен — в .env (JIRA_API_TOKEN)."
        )
    with open(path, "r", encoding="utf-8") as f:
        cfg = json.load(f)
    for key in ("jira_base_url", "jira_email", "month"):
        if key not in cfg or not str(cfg[key]).strip():
            raise ValueError(f"В config.json должно быть заполнено поле: {key}")
    token = os.environ.get("JIRA_API_TOKEN", "").strip() or (cfg.get("jira_api_token") or "").strip()
    if not token:
        raise ValueError(
            "Jira API-токен не задан. Добавьте JIRA_API_TOKEN в файл .env (см. .env.example)."
        )
    cfg["jira_api_token"] = token
    # Нормализуем URL (без слэша в конце)
    cfg["jira_base_url"] = str(cfg["jira_base_url"]).rstrip("/")
    # Проект — обязателен, иначе в выборку попадают задачи из всех проектов
    project = str(cfg.get("project", "")).strip()
    if not project:
        raise ValueError(
            'В config.json должно быть заполнено поле "project" (ключ проекта в Jira, например DB). '
            "Без него в отчёт попадают задачи из всех проектов."
        )
    cfg["project"] = project
    # Проверяем формат месяца YYYY-MM
    month = str(cfg["month"]).strip()
    if not re.match(r"^\d{4}-\d{2}$", month):
        raise ValueError('Поле month должно быть в формате YYYY-MM, например "2025-01"')
    cfg["month"] = month
    return cfg


def month_range(month_yyyy_mm):
    """Возвращает (start_date, end_date_exclusive) для месяца в формате YYYY-MM.
    end_date_exclusive — первый день следующего месяца, для JQL created < end."""
    year, month = int(month_yyyy_mm[:4]), int(month_yyyy_mm[5:7])
    last_day = calendar.monthrange(year, month)[1]
    start = f"{year}-{month:02d}-01"
    # Конец как в фильтре Jira: created < "YYYY-MM-01" следующего месяца
    next_month = month + 1 if month < 12 else 1
    next_year = year if month < 12 else year + 1
    end_exclusive = f"{next_year}-{next_month:02d}-01"
    return start, end_exclusive


def jira_search(auth, base_url, jql, fields=None, max_results=100, fetch_all=True):
    """Поиск задач через Jira REST API (эндпоинт /rest/api/3/search/jql). При fetch_all подгружает все страницы."""
    headers = {"Accept": "application/json"}
    if fields:
        fields_str = ",".join(fields)
    else:
        fields_str = None

    # С мая 2025 старые /rest/api/2/search и /rest/api/3/search возвращают 410 Gone.
    # Используем новый эндпоинт: /rest/api/3/search/jql
    api_path = "/rest/api/3/search/jql"

    def do_page(start_at=0):
        params = {"jql": jql, "maxResults": max_results, "startAt": start_at}
        if fields_str:
            params["fields"] = fields_str
        url = urljoin(base_url + "/", api_path.lstrip("/"))
        r = requests.get(url, auth=auth, headers=headers, params=params, timeout=30)
        r.raise_for_status()
        return r.json()

    data = do_page(0)
    issues = list(data.get("issues", []))
    total = data.get("total", 0)
    while fetch_all and len(issues) < total:
        data = do_page(len(issues))
        issues.extend(data.get("issues", []))
    return issues, None


def extract_issue_row(issue, base_url, is_subtask=False):
    """Из ответа Jira формируем строку для таблицы (ключ, название, статус, дата, ссылка, исполнитель)."""
    key = issue.get("key", "")
    fields = issue.get("fields") or {}
    summary = ""
    if "summary" in fields:
        summary = fields["summary"] or ""
    status = ""
    if "status" in fields and fields["status"]:
        status = fields["status"].get("name") or fields["status"].get("displayName") or ""
    created_raw = fields.get("created") or ""
    created = format_created_dt(parse_jira_dt(created_raw)) if created_raw else ""
    assignee = ""
    if "assignee" in fields and fields["assignee"]:
        assignee = (
            fields["assignee"].get("displayName")
            or fields["assignee"].get("emailAddress")
            or ""
        )
    issue_type = ""
    if "issuetype" in fields and fields["issuetype"]:
        issue_type = (
            fields["issuetype"].get("name") or fields["issuetype"].get("displayName") or ""
        )
    link = f"{base_url}/browse/{key}" if key else ""
    return {
        "key": key,
        "summary": summary,
        "status": status,
        "created": created,
        "assignee": assignee,
        "issuetype": issue_type,
        "link": link,
        "is_subtask": is_subtask,
    }


def parse_jira_dt(dt_str):
    if not dt_str:
        return None
    try:
        return datetime.fromisoformat(dt_str.replace("Z", "+00:00"))
    except ValueError:
        return None


def business_minutes_between(start_dt, end_dt):
    """Рабочее время: 9-12, 13-18, только пн–пт (суббота и воскресенье не считаются)."""
    if not start_dt or not end_dt or end_dt <= start_dt:
        return 0
    total_minutes = 0
    tz = start_dt.tzinfo
    cur_date = start_dt.date()
    end_date = end_dt.date()

    def day_interval(d, t_start, t_end):
        return datetime.combine(d, t_start, tz), datetime.combine(d, t_end, tz)

    while cur_date <= end_date:
        # 0=пн, 5=суббота, 6=воскресенье — не считаем выходные
        if cur_date.weekday() >= 5:
            cur_date += timedelta(days=1)
            continue
        intervals = [
            day_interval(cur_date, WORKDAY_START, WORKDAY_LUNCH_START),
            day_interval(cur_date, WORKDAY_LUNCH_END, WORKDAY_END),
        ]
        for i_start, i_end in intervals:
            s = max(start_dt, i_start)
            e = min(end_dt, i_end)
            if e > s:
                total_minutes += int((e - s).total_seconds() // 60)
        cur_date += timedelta(days=1)
    return total_minutes


def format_work_minutes(minutes):
    if minutes <= 0:
        return "еще не было перехода"
    day_minutes = 8 * 60
    days = minutes // day_minutes
    rem = minutes % day_minutes
    hours = rem // 60
    mins = rem % 60
    parts = []
    if days:
        parts.append(f"{days} дн")
    if hours:
        parts.append(f"{hours} ч")
    if mins or not parts:
        parts.append(f"{mins} мин")
    return " ".join(parts)


def fetch_issue_changelog(auth, base_url, key):
    url = urljoin(base_url + "/", f"/rest/api/3/issue/{key}")
    params = {"expand": "changelog", "fields": "created,status"}
    r = requests.get(url, auth=auth, params=params, timeout=30)
    r.raise_for_status()
    return r.json()


def status_duration_task_to_dev_test(issue_data):
    """Возвращает (рабочие минуты, datetime перехода в DEV TEST) от входа в TASK."""
    fields = issue_data.get("fields") or {}
    created_dt = parse_jira_dt(fields.get("created"))
    current_status = ""
    status_field = fields.get("status") or {}
    if status_field:
        current_status = status_field.get("name") or status_field.get("displayName") or ""

    changelog = (issue_data.get("changelog") or {}).get("histories") or []
    # Сортируем по времени
    changelog_sorted = sorted(changelog, key=lambda h: h.get("created") or "")

    start_dt = None
    end_dt = None

    # Определяем старт по первому переходу или созданию (если стартовый статус = TASK)
    first_status_change = None
    for h in changelog_sorted:
        items = h.get("items") or []
        for it in items:
            if it.get("field") == "status":
                first_status_change = h
                break
        if first_status_change:
            break

    if first_status_change:
        for it in first_status_change.get("items") or []:
            if it.get("field") == "status":
                if it.get("fromString") == STATUS_TASK:
                    start_dt = created_dt
                break

    # Ищем переходы
    for h in changelog_sorted:
        h_dt = parse_jira_dt(h.get("created"))
        for it in h.get("items") or []:
            if it.get("field") != "status":
                continue
            to_status = it.get("toString")
            if start_dt is None and to_status == STATUS_TASK:
                start_dt = h_dt
            elif start_dt is not None and to_status == STATUS_DEV_TEST:
                end_dt = h_dt
                break
        if end_dt:
            break

    # Если нет перехода в DEV TEST
    if not start_dt or not end_dt:
        return 0, None
    return business_minutes_between(start_dt, end_dt), end_dt


def format_dt(dt_obj):
    if not dt_obj:
        return "еще не было перехода"
    return dt_obj.strftime("%d.%m.%Y %H:%M")


def format_created_dt(dt_obj):
    if not dt_obj:
        return ""
    return dt_obj.strftime("%d.%m.%Y %H:%M")


def fetch_buglists_and_subtasks(config):
    """Получить все баглисты за месяц и сабтаски к ним.
    JQL совпадает с фильтром: project = X, issuetype = Баг, created >= start AND created < end.
    В отчёт попадают все задачи типа «Баг» (в т.ч. сабтаски); сабтаски запрашиваются только у родительских задач."""
    auth = HTTPBasicAuth(config["jira_email"], config["jira_api_token"])
    base_url = config["jira_base_url"]
    start, end_exclusive = month_range(config["month"])

    # Базовый JQL как в фильтре Jira: project, тип Баг, даты
    jql_base_parts = [
        f'issuetype = "{ISSUE_TYPE_BUGLIST}"',
        f'created >= "{start}" AND created < "{end_exclusive}"',
    ]
    if config.get("project"):
        jql_base_parts.insert(0, f'project = "{config["project"]}"')
    jql_base = " AND ".join(jql_base_parts)

    fields = ["summary", "status", "created", "assignee", "issuetype", "subtasks", "parent"]
    # Два запроса и объединение: часть API/поиска может не возвращать задачи с родителем (под эпиком).
    # Запрос 1 — без родителя, запрос 2 — с родителем (под эпиком), затем мержим без дубликатов.
    buglists_flat, _ = jira_search(auth, base_url, jql_base + " AND parent is EMPTY", fields=fields)
    buglists_with_parent, _ = jira_search(auth, base_url, jql_base + " AND parent is not EMPTY", fields=fields)
    seen_keys = {i.get("key") for i in buglists_flat if i.get("key")}
    for i in buglists_with_parent:
        if i.get("key") and i.get("key") not in seen_keys:
            seen_keys.add(i.get("key"))
            buglists_flat.append(i)
    buglists = buglists_flat

    result_buglists = []
    result_subtasks = []
    seen_subtask_keys = set()

    for issue in buglists:
        row = extract_issue_row(issue, base_url, is_subtask=False)
        result_buglists.append(row)
        parent_key = issue.get("key")
        # Сабтаски запрашиваем только у родительских задач (у которых нет parent)
        has_parent = bool((issue.get("fields") or {}).get("parent"))
        if has_parent:
            continue

        # Сабтаски: через JQL parent = KEY
        jql_sub = f'parent = "{parent_key}"'
        subs, _ = jira_search(auth, base_url, jql_sub, fields=fields)
        for sub in subs:
            sk = sub.get("key")
            if sk and sk not in seen_subtask_keys:
                seen_subtask_keys.add(sk)
                sr = extract_issue_row(sub, base_url, is_subtask=True)
                sr["parent_key"] = parent_key
                result_subtasks.append(sr)

    return result_buglists, result_subtasks


def fetch_tasks(config):
    """Получить все задачи типа Task за месяц (не сабтаски)."""
    auth = HTTPBasicAuth(config["jira_email"], config["jira_api_token"])
    base_url = config["jira_base_url"]
    start, end_exclusive = month_range(config["month"])

    jql_base_parts = [
        'issuetype = "Task"',
        f'created >= "{start}" AND created < "{end_exclusive}"',
    ]
    if config.get("project"):
        jql_base_parts.insert(0, f'project = "{config["project"]}"')
    jql_base = " AND ".join(jql_base_parts)

    fields = ["summary", "status", "created", "assignee", "issuetype", "parent"]
    tasks_flat, _ = jira_search(auth, base_url, jql_base + " AND parent is EMPTY", fields=fields)
    tasks_with_parent, _ = jira_search(auth, base_url, jql_base + " AND parent is not EMPTY", fields=fields)
    seen_keys = {i.get("key") for i in tasks_flat if i.get("key")}
    for i in tasks_with_parent:
        if i.get("key") and i.get("key") not in seen_keys:
            seen_keys.add(i.get("key"))
            tasks_flat.append(i)

    return [extract_issue_row(t, base_url, is_subtask=False) for t in tasks_flat]


def fetch_misc_subtasks(config):
    """Получить сабтаски за месяц, которые НЕ относятся к баглистам."""
    auth = HTTPBasicAuth(config["jira_email"], config["jira_api_token"])
    base_url = config["jira_base_url"]
    start, end_exclusive = month_range(config["month"])

    jql_base_parts = [
        "issuetype in subtaskIssueTypes()",
        f'created >= "{start}" AND created < "{end_exclusive}"',
    ]
    if config.get("project"):
        jql_base_parts.insert(0, f'project = "{config["project"]}"')
    jql = " AND ".join(jql_base_parts)

    fields = ["summary", "status", "created", "assignee", "issuetype", "parent"]
    subtasks, _ = jira_search(auth, base_url, jql, fields=fields)

    parent_type_cache = {}

    def get_parent_type(parent_key):
        if not parent_key:
            return None
        if parent_key in parent_type_cache:
            return parent_type_cache[parent_key]
        url = urljoin(base_url + "/", f"/rest/api/3/issue/{parent_key}")
        params = {"fields": "issuetype"}
        r = requests.get(url, auth=auth, params=params, timeout=30)
        r.raise_for_status()
        data = r.json()
        p_type = ((data.get("fields") or {}).get("issuetype") or {}).get("name")
        parent_type_cache[parent_key] = p_type
        return p_type

    result = []
    for st in subtasks:
        parent = (st.get("fields") or {}).get("parent") or {}
        parent_key = parent.get("key") or ""
        parent_type = (
            ((parent.get("fields") or {}).get("issuetype") or {}).get("name")
            if isinstance(parent, dict)
            else None
        )
        if not parent_type:
            try:
                parent_type = get_parent_type(parent_key)
            except requests.RequestException:
                parent_type = None
        # Исключаем сабтаски баглистов
        if parent_type == ISSUE_TYPE_BUGLIST:
            continue
        row = extract_issue_row(st, base_url, is_subtask=True)
        row["parent_key"] = parent_key
        row["parent_type"] = parent_type or ""
        result.append(row)
    return result


def write_excel(buglists, subtasks, tasks, misc_subtasks, durations_by_key, devtest_dt_by_key, config, out_path):
    """Пишем два листа: Баглисты и Сабтаски; сводный лист «Сводка»; листы по разработчикам."""
    wb = Workbook()

    # —— Лист «Баглисты» ——
    ws_bug = wb.active
    ws_bug.title = "Баглисты"
    headers_bug = [
        "Ключ",
        "Название",
        "Статус",
        "Создан",
        "Исполнитель",
        f"{STATUS_TASK}→{STATUS_DEV_TEST} (раб. время)",
        f"Дата {STATUS_DEV_TEST}",
        "Ссылка",
    ]
    for c, h in enumerate(headers_bug, 1):
        ws_bug.cell(row=1, column=c, value=h)
    for r, row in enumerate(buglists, 2):
        ws_bug.cell(row=r, column=1, value=row["key"])
        ws_bug.cell(row=r, column=2, value=row["summary"])
        ws_bug.cell(row=r, column=3, value=row["status"])
        ws_bug.cell(row=r, column=4, value=row["created"])
        ws_bug.cell(row=r, column=5, value=row["assignee"])
        ws_bug.cell(row=r, column=6, value=durations_by_key.get(row["key"], "еще не было перехода"))
        ws_bug.cell(row=r, column=7, value=devtest_dt_by_key.get(row["key"], "еще не было перехода"))
        ws_bug.cell(row=r, column=8, value=row["link"])
    for c in range(1, len(headers_bug) + 1):
        ws_bug.cell(row=1, column=c).font = Font(bold=True)
    for col in range(1, 9):
        ws_bug.column_dimensions[get_column_letter(col)].width = 18

    # —— Лист «Сабтаски» ——
    ws_sub = wb.create_sheet("Сабтаски", 1)
    headers_sub = [
        "Родитель (баглист)",
        "Ключ",
        "Название",
        "Статус",
        "Создан",
        "Исполнитель",
        f"{STATUS_TASK}→{STATUS_DEV_TEST} (раб. время)",
        f"Дата {STATUS_DEV_TEST}",
        "Ссылка",
    ]
    for c, h in enumerate(headers_sub, 1):
        ws_sub.cell(row=1, column=c, value=h)
    for r, row in enumerate(subtasks, 2):
        ws_sub.cell(row=r, column=1, value=row.get("parent_key", ""))
        ws_sub.cell(row=r, column=2, value=row["key"])
        ws_sub.cell(row=r, column=3, value=row["summary"])
        ws_sub.cell(row=r, column=4, value=row["status"])
        ws_sub.cell(row=r, column=5, value=row["created"])
        ws_sub.cell(row=r, column=6, value=row["assignee"])
        ws_sub.cell(row=r, column=7, value=durations_by_key.get(row["key"], "еще не было перехода"))
        ws_sub.cell(row=r, column=8, value=devtest_dt_by_key.get(row["key"], "еще не было перехода"))
        ws_sub.cell(row=r, column=9, value=row["link"])
    for c in range(1, len(headers_sub) + 1):
        ws_sub.cell(row=1, column=c).font = Font(bold=True)
    for col in range(1, 10):
        ws_sub.column_dimensions[get_column_letter(col)].width = 20

    # —— Лист «Сводка»: баглист, затем его сабтаски ——
    ws_sum = wb.create_sheet("Сводка", 2)
    headers_sum = [
        "Тип",
        "Родитель",
        "Ключ",
        "Название",
        "Статус",
        "Создан",
        "Исполнитель",
        f"{STATUS_TASK}→{STATUS_DEV_TEST} (раб. время)",
        f"Дата {STATUS_DEV_TEST}",
        "Ссылка",
    ]
    for c, h in enumerate(headers_sum, 1):
        ws_sum.cell(row=1, column=c, value=h)
    for c in range(1, len(headers_sum) + 1):
        ws_sum.cell(row=1, column=c).font = Font(bold=True)
    row_num = 2
    for bl in buglists:
        ws_sum.cell(row=row_num, column=1, value="Баглист")
        ws_sum.cell(row=row_num, column=2, value="")
        ws_sum.cell(row=row_num, column=3, value=bl["key"])
        ws_sum.cell(row=row_num, column=4, value=bl["summary"])
        ws_sum.cell(row=row_num, column=5, value=bl["status"])
        ws_sum.cell(row=row_num, column=6, value=bl["created"])
        ws_sum.cell(row=row_num, column=7, value=bl["assignee"])
        ws_sum.cell(row=row_num, column=8, value=durations_by_key.get(bl["key"], "еще не было перехода"))
        ws_sum.cell(row=row_num, column=9, value=devtest_dt_by_key.get(bl["key"], "еще не было перехода"))
        ws_sum.cell(row=row_num, column=10, value=bl["link"])
        row_num += 1
        for st in subtasks:
            if st.get("parent_key") == bl["key"]:
                ws_sum.cell(row=row_num, column=1, value="Сабтаск")
                ws_sum.cell(row=row_num, column=2, value=bl["key"])
                ws_sum.cell(row=row_num, column=3, value=st["key"])
                ws_sum.cell(row=row_num, column=4, value=st["summary"])
                ws_sum.cell(row=row_num, column=5, value=st["status"])
                ws_sum.cell(row=row_num, column=6, value=st["created"])
                ws_sum.cell(row=row_num, column=7, value=st["assignee"])
                ws_sum.cell(row=row_num, column=8, value=durations_by_key.get(st["key"], "еще не было перехода"))
                ws_sum.cell(row=row_num, column=9, value=devtest_dt_by_key.get(st["key"], "еще не было перехода"))
                ws_sum.cell(row=row_num, column=10, value=st["link"])
                row_num += 1
    for col in range(1, 11):
        ws_sum.column_dimensions[get_column_letter(col)].width = 18

    # —— Листы по разработчикам ——
    def safe_sheet_title(name):
        """Excel ограничивает название листа 31 символом и запрещает спецсимволы."""
        if not name:
            name = "Не назначено"
        name = re.sub(r"[:\\/?*\[\]]+", "_", name).strip()
        return name[:31] if len(name) > 31 else name

    buglist_by_key = {b["key"]: b for b in buglists}
    rows_by_dev = {}

    def add_row(dev, row):
        rows_by_dev.setdefault(dev, []).append(row)

    # Task: если нет исполнителя — в "Не назначено"
    for t in tasks:
        dev = t["assignee"] or "Не назначено"
        add_row(
            dev,
            {
                "section": "TASK",
                "type": "Task",
                "parent_key": "",
                "parent_assignee": "",
                "key": t["key"],
                "summary": t["summary"],
                "status": t["status"],
                "created": t["created"],
                "assignee": t["assignee"],
                "duration": durations_by_key.get(t["key"], "еще не было перехода"),
                "devtest_dt": devtest_dt_by_key.get(t["key"], "еще не было перехода"),
                "link": t["link"],
            },
        )

    # Сабтаски, не относящиеся к баглистам — в раздел задач
    for st in misc_subtasks:
        dev = st["assignee"] or "Не назначено"
        add_row(
            dev,
            {
                "section": "TASK",
                "type": "Сабтаск",
                "parent_key": st.get("parent_key", ""),
                "parent_assignee": "",
                "key": st["key"],
                "summary": st["summary"],
                "status": st["status"],
                "created": st["created"],
                "assignee": st["assignee"],
                "duration": durations_by_key.get(st["key"], "еще не было перехода"),
                "devtest_dt": devtest_dt_by_key.get(st["key"], "еще не было перехода"),
                "link": st["link"],
            },
        )

    # Баглисты: если нет исполнителя — в "Не назначено"
    for bl in buglists:
        dev = bl["assignee"] or "Не назначено"
        add_row(
            dev,
            {
                "section": "BUGS",
                "type": "Баглист",
                "parent_key": "",
                "parent_assignee": "",
                "key": bl["key"],
                "summary": bl["summary"],
                "status": bl["status"],
                "created": bl["created"],
                "assignee": bl["assignee"],
                "duration": durations_by_key.get(bl["key"], "еще не было перехода"),
                "devtest_dt": devtest_dt_by_key.get(bl["key"], "еще не было перехода"),
                "link": bl["link"],
            },
        )

    # Сабтаски: если нет исполнителя — берём исполнителя баглиста,
    # а если у баглиста нет исполнителя — "Не назначено"
    for st in subtasks:
        parent_key = st.get("parent_key", "")
        parent = buglist_by_key.get(parent_key, {})
        parent_assignee = parent.get("assignee") or ""
        assignee = st.get("assignee") or parent_assignee or "Не назначено"
        add_row(
            assignee,
            {
                "section": "BUGS",
                "type": "Сабтаск",
                "parent_key": parent_key,
                "parent_assignee": parent_assignee,
                "key": st["key"],
                "summary": st["summary"],
                "status": st["status"],
                "created": st["created"],
                "assignee": st["assignee"],
                "duration": durations_by_key.get(st["key"], "еще не было перехода"),
                "devtest_dt": devtest_dt_by_key.get(st["key"], "еще не было перехода"),
                "link": st["link"],
            },
        )

    for dev, rows in rows_by_dev.items():
        title = safe_sheet_title(dev)
        # Гарантируем уникальность имени листа
        base = title
        idx = 1
        while title in wb.sheetnames:
            suffix = f"_{idx}"
            title = (base[: 31 - len(suffix)] + suffix) if len(base) + len(suffix) > 31 else base + suffix
            idx += 1
        ws_dev = wb.create_sheet(title)
        headers_dev = [
            "Тип",
            "Баглист",
            "Исполнитель баглиста",
            "Ключ",
            "Название",
            "Статус",
            "Создан",
            "Исполнитель",
            f"{STATUS_TASK}→{STATUS_DEV_TEST} (раб. время)",
            f"Дата {STATUS_DEV_TEST}",
            "Ссылка",
        ]
        for c, h in enumerate(headers_dev, 1):
            ws_dev.cell(row=1, column=c, value=h).font = Font(bold=True)
        row_num = 2
        tasks_rows = [r for r in rows if r.get("section") == "TASK"]
        task_sub_rows = [r for r in tasks_rows if r.get("type") == "Сабтаск"]
        task_main_rows = [r for r in tasks_rows if r.get("type") != "Сабтаск"]
        bug_rows = [r for r in rows if r.get("section") == "BUGS"]

        if tasks_rows:
            ws_dev.cell(row=row_num, column=1, value="Задачи (Task)").font = Font(bold=True)
            row_num += 1
            for row in task_main_rows:
                ws_dev.cell(row=row_num, column=1, value=row["type"])
                ws_dev.cell(row=row_num, column=2, value=row["parent_key"])
                ws_dev.cell(row=row_num, column=3, value=row["parent_assignee"])
                ws_dev.cell(row=row_num, column=4, value=row["key"])
                ws_dev.cell(row=row_num, column=5, value=row["summary"])
                ws_dev.cell(row=row_num, column=6, value=row["status"])
                ws_dev.cell(row=row_num, column=7, value=row["created"])
                ws_dev.cell(row=row_num, column=8, value=row["assignee"])
                ws_dev.cell(row=row_num, column=9, value=row["duration"])
                ws_dev.cell(row=row_num, column=10, value=row["devtest_dt"])
                ws_dev.cell(row=row_num, column=11, value=row["link"])
                row_num += 1
            if task_sub_rows:
                ws_dev.cell(row=row_num, column=1, value="Сабтаски (не баглисты)").font = Font(bold=True)
                row_num += 1
                for row in task_sub_rows:
                    ws_dev.cell(row=row_num, column=1, value=row["type"])
                    ws_dev.cell(row=row_num, column=2, value=row["parent_key"])
                    ws_dev.cell(row=row_num, column=3, value=row["parent_assignee"])
                    ws_dev.cell(row=row_num, column=4, value=row["key"])
                    ws_dev.cell(row=row_num, column=5, value=row["summary"])
                    ws_dev.cell(row=row_num, column=6, value=row["status"])
                    ws_dev.cell(row=row_num, column=7, value=row["created"])
                    ws_dev.cell(row=row_num, column=8, value=row["assignee"])
                    ws_dev.cell(row=row_num, column=9, value=row["duration"])
                    ws_dev.cell(row=row_num, column=10, value=row["devtest_dt"])
                    ws_dev.cell(row=row_num, column=11, value=row["link"])
                    row_num += 1

        if bug_rows:
            if row_num > 2:
                row_num += 1
            ws_dev.cell(row=row_num, column=1, value="Баглисты и сабтаски").font = Font(bold=True)
            row_num += 1
            for row in bug_rows:
                ws_dev.cell(row=row_num, column=1, value=row["type"])
                ws_dev.cell(row=row_num, column=2, value=row["parent_key"])
                ws_dev.cell(row=row_num, column=3, value=row["parent_assignee"])
                ws_dev.cell(row=row_num, column=4, value=row["key"])
                ws_dev.cell(row=row_num, column=5, value=row["summary"])
                ws_dev.cell(row=row_num, column=6, value=row["status"])
                ws_dev.cell(row=row_num, column=7, value=row["created"])
                ws_dev.cell(row=row_num, column=8, value=row["assignee"])
                ws_dev.cell(row=row_num, column=9, value=row["duration"])
                ws_dev.cell(row=row_num, column=10, value=row["devtest_dt"])
                ws_dev.cell(row=row_num, column=11, value=row["link"])
                row_num += 1
        for col in range(1, 12):
            ws_dev.column_dimensions[get_column_letter(col)].width = 20

    wb.save(out_path)
    return out_path


def main():
    config = load_config()
    print(f"Месяц из конфига: {config['month']}")
    print("Запрос баглистов (тип «Баг») и сабтасков...")
    buglists, subtasks = fetch_buglists_and_subtasks(config)
    print('Запрос задач типа "Task"...')
    tasks = fetch_tasks(config)
    print("Запрос сабтасков, не относящихся к баглистам...")
    misc_subtasks = fetch_misc_subtasks(config)
    print(f"Найдено баглистов: {len(buglists)}, сабтасков: {len(subtasks)}")
    print(f'Найдено задач типа "Task": {len(tasks)}')
    print(f"Найдено сабтасков (не баглисты): {len(misc_subtasks)}")

    print(f'Подсчёт рабочего времени перехода {STATUS_TASK} → {STATUS_DEV_TEST}...')
    auth = HTTPBasicAuth(config["jira_email"], config["jira_api_token"])
    base_url = config["jira_base_url"]
    all_keys = [r["key"] for r in (buglists + subtasks + tasks + misc_subtasks)]
    durations_by_key = {}
    devtest_dt_by_key = {}
    unique_keys = sorted(set(all_keys))

    def process_key(k):
        issue_data = fetch_issue_changelog(auth, base_url, k)
        minutes, end_dt = status_duration_task_to_dev_test(issue_data)
        return k, format_work_minutes(minutes), format_dt(end_dt)

    with ThreadPoolExecutor(max_workers=8) as executor:
        futures = {executor.submit(process_key, k): k for k in unique_keys}
        for fut in as_completed(futures):
            key = futures[fut]
            try:
                k, dur, dt_val = fut.result()
                durations_by_key[k] = dur
                devtest_dt_by_key[k] = dt_val
            except requests.RequestException:
                durations_by_key[key] = "еще не было перехода"
                devtest_dt_by_key[key] = "еще не было перехода"
            except Exception:
                durations_by_key[key] = "еще не было перехода"
                devtest_dt_by_key[key] = "еще не было перехода"

    out_name = f"jira_buglist_report_{config['month']}.xlsx"
    out_path = os.path.join(SCRIPT_DIR, out_name)
    write_excel(buglists, subtasks, tasks, misc_subtasks, durations_by_key, devtest_dt_by_key, config, out_path)
    print(f"Отчёт сохранён: {out_path}")


if __name__ == "__main__":
    main()
